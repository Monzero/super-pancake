import pytest

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl is required to run workbook tests")
from openpyxl import Workbook

from workbook_validator import WorkbookValidationError, validate_workbook


def _create_workbook(schema_headers=None, data_header=None):
    wb = Workbook()
    schema_ws = wb.active
    schema_ws.title = "Schema"
    if schema_headers is None:
        schema_headers = ["field_name", "data_type", "nullable"]
    schema_ws.append(schema_headers)
    schema_ws.append(["id", "integer", False])

    data_ws = wb.create_sheet("Data")
    if data_header is None:
        data_header = ["id"]
    data_ws.append(data_header)
    return wb


def test_valid_workbook(tmp_path):
    wb = _create_workbook()
    path = tmp_path / "valid.xlsx"
    wb.save(path)
    validate_workbook(path)


def test_missing_required_sheet(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Schema"
    ws.append(["field_name", "data_type", "nullable"])
    ws.append(["id", "integer", False])
    path = tmp_path / "missing_data.xlsx"
    wb.save(path)
    with pytest.raises(WorkbookValidationError):
        validate_workbook(path)


def test_mismatched_data_headers(tmp_path):
    wb = _create_workbook(data_header=["wrong"])
    path = tmp_path / "mismatch.xlsx"
    wb.save(path)
    with pytest.raises(WorkbookValidationError):
        validate_workbook(path)
