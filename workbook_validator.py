"""Workbook validation utilities."""
from __future__ import annotations

from typing import List

from openpyxl import load_workbook


REQUIRED_SHEETS = ["Schema", "Data"]
REQUIRED_SCHEMA_COLUMNS = ["field_name", "data_type", "nullable"]


class WorkbookValidationError(Exception):
    """Raised when a workbook fails validation."""


def _get_header(row) -> List[str]:
    """Return trimmed string values from a row of cells."""
    values: List[str] = []
    for cell in row:
        value = cell.value
        if isinstance(value, str):
            value = value.strip()
        values.append(value)
    return values


def validate_workbook(path: str) -> None:
    """Validate that the workbook at *path* meets requirements.

    The function raises :class:`WorkbookValidationError` with a descriptive
    message when validation fails. Successful validation returns ``None``.
    """
    try:
        wb = load_workbook(filename=path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - exceptional path
        raise WorkbookValidationError(f"Unable to open workbook: {exc}") from exc

    try:
        # Detect missing required sheets
        missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
        if missing:
            print(f"Missing required sheet(s): {missing}")
            raise WorkbookValidationError(
                f"Workbook missing required sheet(s): {', '.join(missing)}"
            )

        schema_ws = wb["Schema"]
        # Retrieve header row from Schema sheet
        header = _get_header(next(schema_ws.iter_rows(min_row=1, max_row=1)))
        # Detect missing required schema columns
        missing_cols = [c for c in REQUIRED_SCHEMA_COLUMNS if c not in header]
        if missing_cols:
            print(f"Missing required schema column(s): {missing_cols}")
            raise WorkbookValidationError(
                "Schema sheet missing required column(s): "
                + ", ".join(missing_cols)
            )

        # Obtain declared field names from schema
        field_names = [
            row[0].value
            for row in schema_ws.iter_rows(min_row=2, max_col=1)
            if row[0].value is not None
        ]

        data_ws = wb["Data"]
        # Retrieve header row from Data sheet
        data_header = _get_header(next(data_ws.iter_rows(min_row=1, max_row=1)))
        # Compare Data sheet headers with schema field names
        if data_header != field_names:
            print(f"Schema field names: {field_names}")
            print(f"Data sheet headers: {data_header}")
            raise WorkbookValidationError(
                "Data sheet headers do not match field names defined in Schema"
            )
    finally:
        wb.close()
